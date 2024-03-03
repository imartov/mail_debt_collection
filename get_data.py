'''
This module contains methods for retrieving data from sources (export data methods)
'''

import os
import logging
from dotenv import load_dotenv
from openpyxl import load_workbook


logging.basicConfig(level=logging.DEBUG, filename="py_log.log",filemode="w",
                    format="%(asctime)s %(levelname)s %(message)s")
load_dotenv()

def get_data() -> list:
    ''' this method retrieves data from the source file '''
    wb = load_workbook(filename=os.getenv("SOURCE_FILE"))
    ws = wb.active
    recipients = []
    for row in ws.iter_rows(min_row=3):
        unp = row[11].value
        monolit_code = row[0].value
        client_name = row[1].value
        debt_amount = row[4].value
        payment_date = row[12].value
        email = row[13].value

        sfilter = ServiceFilter(unp=unp, client_name=client_name, debt_sum=debt_amount,
                                payment_date=payment_date, email=email)
        if not sfilter.run():
            continue

        recipients.append(
            {
                "unp": unp,
                "monolit_code": monolit_code,
                "client_name": client_name,
                "debt_amount": debt_amount,
                "payment_date": payment_date,
                "email": email
            }
        )
    return recipients

class ServiceFilter:
    ''' this class contains methods for checking values and filtering input data '''
    def __init__(self, **kwargs) -> None:
        self.kwargs = kwargs

    def check_exist_email(self):
        ''' this method checks for the presence of the client's email address '''
        email = self.kwargs.get('email')
        if email:
            return True
        return None

    def check_exist_debt_amount(self):
        ''' this method checks for overdue customer receivables'''
        debt_sum = self.kwargs.get('debt_sum')
        if debt_sum:
            return True
        return None

    def run(self) -> None:
        ''' this method determines the order in which class methods are called '''
        if self.check_exist_email() and self.check_exist_debt_amount():
            return True
        return None


def main():
    ''' this method controls calls to other module methods '''
    get_data()

if __name__ == "__main__":
    main()
